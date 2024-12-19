import sys
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget, QLineEdit, QLabel,
    QComboBox, QMessageBox, QCheckBox, QFormLayout, QHBoxLayout, QStackedWidget
)
from PyQt5.QtGui import QIntValidator
from PyQt5.QtGui import QRegularExpressionValidator
from PyQt5.QtCore import QRegularExpression
from openpyxl import Workbook, load_workbook
import os

# Path to Excel file
EXCEL_FILE = "\survey_data.xlsx"

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Главное окно")
        self.setGeometry(100, 100, 300, 200)

        self.button = QPushButton("Создать запись")
        self.button.clicked.connect(self.open_survey_form)

        layout = QVBoxLayout()
        layout.addWidget(self.button)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def open_survey_form(self):
        self.survey_form = SurveyForm()
        self.survey_form.show()
        self.close()

class SurveyForm(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Опросник")
        self.setGeometry(100, 100, 400, 600)

        self.is_empty = False
        self.gender_m = True
        self.gender_f = False

        self.layout = QVBoxLayout()
        self.form_layout = QFormLayout()

        # User info fields
        self.snils_input = QLineEdit()
        self.snils_input.setPlaceholderText("XXX-XXX-XXX YY")
        self.snils_input.setInputMask("000-000-000 00")

        self.fio_input = QLineEdit()
        regex = QRegularExpression("[А-Яа-яЁё\s-]*") 
        validator = QRegularExpressionValidator(regex, self.fio_input)
        self.fio_input.setValidator(validator)

        self.phone_input = QLineEdit()
        self.phone_input.setPlaceholderText("X (XXX) XXX-XX-XX")
        self.phone_input.setInputMask("0 (000) 000-00-00")

        self.age_input = QLineEdit()
        self.age_input.setValidator(QIntValidator(50, 99, self))
        self.age_input.editingFinished.connect(self.check_value)
        self.age_input.setFocus()

        self.weight_input = QLineEdit()
        self.weight_input.setValidator(QIntValidator(38, 200, self))

        self.gender_input = QComboBox()
        self.gender_input.addItems(["М", "Ж"])
        self.gender_input.currentIndexChanged.connect(self.on_combobox_changed)

        self.height_input = QLineEdit()
        self.height_input.setValidator(QIntValidator(120, 210, self))

        self.region = QComboBox()
        self.region.addItems(["Бурятия", "Татарстан", "Кабардино-Балкария", "Челябинская область", "Томская область", "Рязанская область"])

        # Add fields to layout
        self.form_layout.addRow("Регион:", self.region)
        self.form_layout.addRow("СНИЛС:", self.snils_input)
        self.form_layout.addRow("ФИО:", self.fio_input)
        self.form_layout.addRow("Телефон:", self.phone_input)
        self.form_layout.addRow("Возраст:", self.age_input)
        self.form_layout.addRow("Пол:", self.gender_input)
        self.form_layout.addRow("Рост:", self.height_input)
        self.form_layout.addRow("Вес:", self.weight_input)

        self.layout.addLayout(self.form_layout)

        # Create stacked widget for questions
        self.stacked_widget = QStackedWidget()

        self.questions_layout = QVBoxLayout()
        self.first_question = QComboBox()
        self.first_question.addItems(["Очень светлая кожа, есть веснушки, быстро «сгораю» на солнце", "Светлая кожа цвета «слоновой кости», медленно загораю, но кожа легко краснеет на солнце", 
                                      "Светлая кожа цвета «слоновой кости», загар проявляется постепенно и равномерно", "Кожа оливкового оттенка, практически не обгорает, загар ложится равномерно и интенсивно ", 
                                      "Смуглая или темно-коричневая кожа, редко обгорает, загар проявляется насыщенным темным оттенком", "Очень смуглая кожа шоколадного оттенка с низкой чувствительностью к ультрафиолету: никогда не обгорает, загар очень темный"])
        self.questions_layout.addWidget(QLabel("2. Как бы Вы охарактеризовали свой тип кожи?"))
        self.questions_layout.addWidget(self.first_question)
       
        self.questions = []
        self.list_of_texts = []

        text = "1. Курите ли Вы в настоящее время?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        self.questions.append(self.questions_layout)
        self.list_of_texts.append("2. Как бы Вы охарактеризовали свой тип кожи?")
        text = "3. Используете ли Вы косметические средства (например, крема, спреи) с солнцезащитными факторами (SPF) на регулярной основе?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget_with_only(text))
        text = "4. Сколько минут в день в среднем вы проводите на улице?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_sun_question_widget(text))
        text = "5. Посещаете ли Вы солярий?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "6. Употребляете ли Вы в настоящее время более 1 бокала вина (120 мл) или 1 бокала пива (285 г) или 30 мл крепкого спиртного в день?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "7. Присутствуют ли в Вашем рационе цельное молоко – минимум 1 порция в сутки?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "8. Присутствуют ли в Вашем рационе иные молочные продукты – минимум 2 порции в неделю?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "9. Присутствуют ли в Вашем рационе рыба – минимум 3 порции в неделю?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "10. Присутствуют ли в Вашем рационе яйца – минимум 2 штуки в неделю?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "11. Уменьшился ли Ваш рост более чем на 4 см за жизнь?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "12. Уменьшился ли Ваш рост более чем на 2 см за прошедший год?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "13. Были ли у Вас во взрослом возрасте переломы костей, произошедшие самопроизвольно или в результате такой травмы, от которой у обычного человека перелом бы не возник?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "14. Был ли у кого-то из Ваших родителей перелом шейки бедра?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget_with_idk(text))
        text = "15. Беспокоят ли Вас жажда, частое ночное мочеиспускание?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "16. Были ли у Вас в течение жизни язвы и/или эрозии желудка или двенадцатиперстной кишки?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget_with_idk(text))
        text = "17. Обнаруживались ли у Вас когда-либо камни или «песок» в почках? Были ли почечные колики?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget_with_idk(text))
        text = "18. Были ли у Вас операции на органах шеи?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "19. Страдаете ли Вы или страдали ранее какими-либо из нижеперечисленных заболеваний?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_simple_question_widget(text))
        text = "Ревматоидный артрит"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Болезнь Крона, язвенный колит, целиакия"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Сахарный диабет"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Тиреотоксикоз, гипертиреоз"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Рак простаты (для мужчин)"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Рак молочной железы (для женщин)"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Менопауза до 45 лет (в том числе, операции по удалению яичников) (для женщин)"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Гипогонадизм (для мужчин)"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Хроническое заболевание печени"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Хроническое заболевание почек"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Хроническая обструктивная болезнь легких или иное хроническое заболевание легких"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "ВИЧ"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Вынужденное снижение двигательной активности"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "20. Принимаете ли Вы в настоящее время или принимали ли Вы когда-либо глюкокортикоиды («стероиды») в течение 3 месяцев и более?"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget_with_idk(text))
        text = "21. Принимаете ли Вы в настоящее время или принимали в течение последнего года препараты из этих групп:"
        self.list_of_texts.append(text)
        self.questions.append(self.create_simple_question_widget(text))
        text = "Антиконвульсанты (вальпроат натрия, карбамазепин, этосуксимид, ламотриджин, топирамат, леветирацетам и др.)"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Антипсихотики (галоперидол, арипипразол, брекспипразол, кветиапин, клозапин, оланзапин, амисульприд, перфеназин, хлорпромазин, хлорпротиксен и др.)"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Антиретровирусные препараты (зиновудин, ламивудин, тенофовир, фосфазид, эфавиренз, этравирин, индинавир, фосампренавир, дарунавир, абакавир, эмтрицитабин, лопинавир, ритонавир, фостемсавир, маравирок и др.)"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Препараты для снижения массы тела (орлистат, сибутрамин, лираглутид, семаглутид)"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Препараты для снижения холестерина (статины, эзетимиб, эволокумаб)"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Слабительные"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))
        text = "Лекарственные препараты и/или БАДы с витамином D"
        self.list_of_texts.append(text)
        self.questions.append(self.create_question_widget(text))

        self.questions[24].itemAt(1).widget().setDisabled(True)
        self.questions[24].itemAt(2).widget().setDisabled(True)
        self.questions[25].itemAt(1).widget().setDisabled(True)
        self.questions[25].itemAt(2).widget().setDisabled(True)
        self.questions[23].itemAt(1).widget().setDisabled(False)
        self.questions[23].itemAt(2).widget().setDisabled(False) 
        self.questions[26].itemAt(1).widget().setDisabled(False)
        self.questions[26].itemAt(2).widget().setDisabled(False)
        

        # Create pages with 5 questions each
        for i in range(0, len(self.questions), 6):
            page_layout = QVBoxLayout()
            for j in range(i, min(i + 6, len(self.questions))):
                page_layout.addLayout(self.questions[j])

            page_widget = QWidget()
            page_widget.setLayout(page_layout)
            self.stacked_widget.addWidget(page_widget)

        self.layout.addWidget(self.stacked_widget)

        # Navigation buttons
        self.nav_layout = QHBoxLayout()
        self.prev_button = QPushButton("Назад")
        self.next_button = QPushButton("Далее")
        self.prev_button.clicked.connect(self.go_to_previous_page)
        self.next_button.clicked.connect(self.go_to_next_page)

        self.nav_layout.addWidget(self.prev_button)
        self.nav_layout.addWidget(self.next_button)
        self.layout.addLayout(self.nav_layout)

        # Submit button (only visible on last page)
        self.submit_button = QPushButton("Выгрузить в Excel")
        self.submit_button.clicked.connect(self.export_to_excel)
        self.layout.addWidget(self.submit_button)
        self.submit_button.hide()

        self.setLayout(self.layout)
        self.update_navigation()

    def focusOutEvent(self, event):
        text = self.age_input.text()
        if text:
            value = int(text)
            if value < 50 or value > 99:
                self.age_input.setText('50')
        super().focusOutEvent(event)

    def check_value(self):
        text = self.age_input.text()
        if text:
            value = int(text)
            if value < 50 or value > 99:
                self.age_input.setText('50')

    def create_simple_question_widget(self, question_text):
        question_label = QLabel(question_text)
        question_layout = QVBoxLayout()
        question_layout.addWidget(question_label)
        return question_layout
    
    def create_sun_question_widget(self, question_text):
        question_label = QLabel(question_text)
        question_layout = QVBoxLayout()
        min_input = QLineEdit()
        min_input.setValidator(QIntValidator(0, 960, self))
        question_layout.addWidget(question_label)
        question_layout.addWidget(min_input)
        return question_layout

    def create_question_widget(self, question_text):
        question_label = QLabel(question_text)
        yes_checkbox = QCheckBox("Да")
        no_checkbox = QCheckBox("Нет")

        # Ensure mutual exclusivity
        yes_checkbox.stateChanged.connect(lambda state, n=no_checkbox: n.setChecked(False))
        no_checkbox.stateChanged.connect(lambda state, y=yes_checkbox: y.setChecked(False))

        question_layout = QVBoxLayout()
        question_layout.addWidget(question_label)
        question_layout.addWidget(yes_checkbox)
        question_layout.addWidget(no_checkbox)

        return question_layout
    
    def create_question_widget_with_idk(self, question_text):
        question_label = QLabel(question_text)
        yes_checkbox = QCheckBox("Да")
        no_checkbox = QCheckBox("Нет")
        idk_checkbox = QCheckBox("Не знаю")

        # Ensure mutual exclusivity
        yes_checkbox.stateChanged.connect(lambda state, n=no_checkbox: n.setChecked(False))
        yes_checkbox.stateChanged.connect(lambda state, idk=idk_checkbox: idk.setChecked(False))
        no_checkbox.stateChanged.connect(lambda state, y=yes_checkbox: y.setChecked(False))
        no_checkbox.stateChanged.connect(lambda state, idk=idk_checkbox: idk.setChecked(False))
        idk_checkbox.stateChanged.connect(lambda state, n=no_checkbox: n.setChecked(False))
        idk_checkbox.stateChanged.connect(lambda state, y=yes_checkbox: y.setChecked(False))

        question_layout = QVBoxLayout()
        question_layout.addWidget(question_label)
        question_layout.addWidget(yes_checkbox)
        question_layout.addWidget(no_checkbox)
        question_layout.addWidget(idk_checkbox)

        return question_layout
    
    def create_question_widget_with_only(self, question_text):
        question_label = QLabel(question_text)
        yes_checkbox = QCheckBox("Да")
        no_checkbox = QCheckBox("Нет")
        only_checkbox = QCheckBox("Только летом")

        # Ensure mutual exclusivity
        yes_checkbox.stateChanged.connect(lambda state, n=no_checkbox: n.setChecked(False))
        yes_checkbox.stateChanged.connect(lambda state, idk=only_checkbox: idk.setChecked(False))
        no_checkbox.stateChanged.connect(lambda state, y=yes_checkbox: y.setChecked(False))
        no_checkbox.stateChanged.connect(lambda state, idk=only_checkbox: idk.setChecked(False))
        only_checkbox.stateChanged.connect(lambda state, n=no_checkbox: n.setChecked(False))
        only_checkbox.stateChanged.connect(lambda state, y=yes_checkbox: y.setChecked(False))

        question_layout = QVBoxLayout()
        question_layout.addWidget(question_label)
        question_layout.addWidget(yes_checkbox)
        question_layout.addWidget(no_checkbox)
        question_layout.addWidget(only_checkbox)

        return question_layout

    def go_to_previous_page(self):
        current_index = self.stacked_widget.currentIndex()
        if current_index > 0:
            self.stacked_widget.setCurrentIndex(current_index - 1)
        self.update_navigation()

    def go_to_next_page(self):
        current_index = self.stacked_widget.currentIndex()
        if current_index < self.stacked_widget.count() - 1:
            self.stacked_widget.setCurrentIndex(current_index + 1)
        self.update_navigation()

    def update_navigation(self):
        current_index = self.stacked_widget.currentIndex()
        self.prev_button.setEnabled(current_index > 0)
        self.next_button.setEnabled(current_index < self.stacked_widget.count() - 1)
        self.submit_button.setVisible(current_index == self.stacked_widget.count() - 1)

    def check_empty(self):
         QMessageBox.information(self, "Warning!", "Заполните все поля!")
         self.is_empty = True
    
    def on_combobox_changed(self):
        if self.gender_input.currentText() == "Ж":
            self.questions[23].itemAt(1).widget().setDisabled(True)
            self.questions[23].itemAt(2).widget().setDisabled(True) 
            self.questions[26].itemAt(1).widget().setDisabled(True)
            self.questions[26].itemAt(2).widget().setDisabled(True)
            self.questions[24].itemAt(1).widget().setDisabled(False) 
            self.questions[24].itemAt(2).widget().setDisabled(False)
            self.questions[25].itemAt(1).widget().setDisabled(False)
            self.questions[25].itemAt(2).widget().setDisabled(False)
            self.gender_f = True
            self.gender_m = False
        else:
            self.questions[24].itemAt(1).widget().setDisabled(True)
            self.questions[24].itemAt(2).widget().setDisabled(True)
            self.questions[25].itemAt(1).widget().setDisabled(True)
            self.questions[25].itemAt(2).widget().setDisabled(True)
            self.questions[23].itemAt(1).widget().setDisabled(False)
            self.questions[23].itemAt(2).widget().setDisabled(False) 
            self.questions[26].itemAt(1).widget().setDisabled(False)
            self.questions[26].itemAt(2).widget().setDisabled(False)
            self.gender_f = False
            self.gender_m = True
        

    def export_to_excel(self):
        # Collect user data
        if self.snils_input.text():
            if self.fio_input.text():
                if self.phone_input.text():
                    if self.age_input.text():
                        if self.height_input.text():
                            if self.weight_input.text():
                                user_data = [
                                    self.region.currentText(),
                                    self.snils_input.text(),
                                    self.fio_input.text(),
                                    self.phone_input.text(),
                                    self.age_input.text(),
                                    self.gender_input.currentText(),
                                    self.height_input.text(),
                                    self.weight_input.text()
                                ]
                            else:
                                self.check_empty()
                                if self.is_empty:
                                    self.is_empty = False
                                    return
                        else:
                                self.check_empty()
                                if self.is_empty:
                                    self.is_empty = False
                                    return
                    else:
                                self.check_empty()
                                if self.is_empty:
                                    self.is_empty = False
                                    return
                else:
                                self.check_empty()
                                if self.is_empty:
                                    self.is_empty = False
                                    return
            else:
                                self.check_empty()
                                if self.is_empty:
                                    self.is_empty = False
                                    return
        else:
                                self.check_empty()
                                if self.is_empty:
                                    self.is_empty = False
                                    return

        # Collect responses
        question_data = []
        counter = 0
        for question_layout in self.questions:
            counter = counter + 1
            try:
                yes_checkbox = question_layout.itemAt(1).widget()
                no_checkbox = question_layout.itemAt(2).widget()
                idk_checkbox = question_layout.itemAt(3).widget()
                if idk_checkbox.text() == "Только летом":
                    question_data.append("Да" if yes_checkbox.isChecked() else "Нет" if no_checkbox.isChecked() else "Только летом" if idk_checkbox.isChecked() else self.check_empty())
                    if self.is_empty:
                        self.is_empty = False
                        return
                else:
                    question_data.append("Да" if yes_checkbox.isChecked() else "Нет" if no_checkbox.isChecked() else "Не знаю" if idk_checkbox.isChecked() else self.check_empty())
                    if self.is_empty:
                        self.is_empty = False
                        return
            except:
                try:
                    if counter == 25 or counter == 26 or counter == 24 or counter == 27:
                        if self.gender_m:
                            if counter == 25 or counter == 26:
                                question_data.append("")
                            else:
                                yes_checkbox = question_layout.itemAt(1).widget()
                                no_checkbox = question_layout.itemAt(2).widget()
                                question_data.append("Да" if yes_checkbox.isChecked() else "Нет" if no_checkbox.isChecked() else self.check_empty())
                                if self.is_empty:
                                    self.is_empty = False
                                    return                            
                        if self.gender_f:
                            if counter == 24 or counter == 27:
                                question_data.append("")
                            else:
                                yes_checkbox = question_layout.itemAt(1).widget()
                                no_checkbox = question_layout.itemAt(2).widget()
                                question_data.append("Да" if yes_checkbox.isChecked() else "Нет" if no_checkbox.isChecked() else self.check_empty())
                                if self.is_empty:
                                    self.is_empty = False
                                    return
                    else:                           
                        yes_checkbox = question_layout.itemAt(1).widget()
                        no_checkbox = question_layout.itemAt(2).widget()
                        question_data.append("Да" if yes_checkbox.isChecked() else "Нет" if no_checkbox.isChecked() else self.check_empty())
                        if self.is_empty:
                            self.is_empty = False
                            return
                except:
                    try:
                        yes_checkbox = question_layout.itemAt(1).widget()
                        question_data.append(yes_checkbox.currentText())
                    except:
                        try:
                            yes_checkbox = question_layout.itemAt(1).widget()
                            if yes_checkbox.text():
                                question_data.append(yes_checkbox.text())
                            else:
                                self.check_empty()
                                if self.is_empty:
                                    self.is_empty = False
                                    return
                        except:                        
                            question_data.append("")
        counter = 0
            

        # Save to Excel
        num_rows = ["1"]
        full_path = os.path.dirname(os.path.abspath(__file__))
        
        if os.path.exists(full_path + EXCEL_FILE):
            workbook = load_workbook(full_path + EXCEL_FILE)
            sheet = workbook.active
            num_rows[0] = sheet.max_row
        else:
            workbook = Workbook()
            sheet = workbook.active
            # Create headers
            headers = ["№", "Регион", "СНИЛС", "ФИО", "Телефон", "Возраст", "Пол", "Рост", "Вес"] + [i for i in self.list_of_texts]
            sheet.append(headers)
        
        
        sheet.append(num_rows + user_data + question_data)
        try:
            workbook.save(full_path + EXCEL_FILE)
            QMessageBox.information(self, "Успех", f"Данные успешно выгружены в Excel! {full_path}{EXCEL_FILE}")
            self.return_to_main()
        except:
            QMessageBox.information(self, "Warning!", "Перед сохранением закройте файл Excel!")
           
    def return_to_main(self):
        self.main_window = MainWindow()
        self.main_window.show()
        self.close()

if __name__ == "__main__":
    app = QApplication(sys.argv)

    main_window = MainWindow()
    main_window.show()

    sys.exit(app.exec_())